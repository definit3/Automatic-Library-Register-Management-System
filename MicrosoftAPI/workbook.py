import openpyxl
from openpyxl import Workbook
from openpyxl.cell import  Cell


wb=Workbook()

ws=wb.active
ws=wb.create_sheet("Register",0)
ws.cell(row=1,column=1,value='Roll')
ws.cell(row=1,column=2,value='Name')
ws.cell(row=1,column=3,value='Date')
ws.cell(row=1,column=4,value='In Time')
ws.cell(row=1,column=5,value='Out Time')

wb.save(filename='workbook.xlsx')

