import cv2
import dlib
import os
import sys
import sqlite3
import cognitive_face as CF
from global_variables import personGroupId
import urllib
from openpyxl import Workbook, load_workbook
from openpyxl.cell import  Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import datetime
import urllib3
urllib3.disable_warnings()
import numpy as np

connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

wb = load_workbook(filename = "workbook.xlsx")
ws=wb.active
sheet = wb.get_sheet_by_name('Register')


Key = 'db85583701c04bf08f8a9fb9a9290b78'
CF.Key.set(Key)
BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0'
CF.BaseUrl.set(BASE_URL)

cap = cv2.VideoCapture(0)
#address = "http://172.16.181.195:8080/video" 
#cap.open(address)
cv2.namedWindow("test")
img_counter = 0
detector = dlib.get_frontal_face_detector()

def blankrow():
	for i in range(1, 100000):
		r=ws.cell(row=i,column=1).value
		if r is None:
			return i
	else:
		 print 'no blank index found'


def findrow():
	for i in range(1,x):
		if ws.cell(row=i,column=1).value==row[2]:
			if ws.cell(row=i,column=5).value is None:
				return i
	else :
		p=0
		return p

			

while(True):

   	 ret, img = cap.read()
   	 cv2.imshow("test", img)
   	 if not ret:
   	     break
   	 k = cv2.waitKey(1)

   	 if k%256 == 27:
   	     # ESC pressed
   	     print("Escape hit, closing...")
   	     break
	 elif k==-1:
	     continue

         elif k%256 == 32:
   	     # SPACE pressed
	     dets = detector(img, 1)

	     if not os.path.exists('./detected_face'):
		os.makedirs('./detected_face')
	     print "detected = " + str(len(dets))
	     if(len(dets)==0):
	     	continue
	     for i, d in enumerate(dets):
	     	cv2.imwrite('./detected_face/face' + str(i + 1) + '.jpg', img[d.top():d.bottom(), d.left():d.right()])

   	     print("{} written!")
   	     img_counter += 1

                  
         currentDir = os.path.dirname(os.path.abspath(__file__))
	 directory = os.path.join(currentDir, 'detected_face')
	 for filename in os.listdir(directory):
		if filename.endswith(".jpg"):
			imgurl = urllib.pathname2url(os.path.join(directory, filename))
			res = CF.face.detect(imgurl)
			if len(res) != 1:
				print "No face detected."
				continue
				
			faceIds = []
			for face in res:
				faceIds.append(face['faceId'])
			res = CF.face.identify(faceIds, personGroupId)
			print filename
			print res
			for face  in res:
				if not face['candidates']:
					print "Unknown"
				else:
					personId = face['candidates'][0]['personId']
					c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
					row = c.fetchone()
					print row[1] + " recognized"

					x=blankrow()
					y=findrow()
					if y:
						time=datetime.datetime.now().time()
						ws.cell(row=y,column=5,value=time)
						wb.save(filename='workbook.xlsx')
					else :
						print 'not done'
					
					
			
cap.release()                                                                  
cv2.destroyAllWindows()




